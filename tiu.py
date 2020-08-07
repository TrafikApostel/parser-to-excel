import requests
from openpyxl import Workbook

i = 2
wb = Workbook()
ws = wb.active
ws['A1'] = 'text'
header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36 OPR/68.0.3618.206 (Edition Yx GX)'
}
urls = ['https://zakaz.atbmarket.com/', 'https://ta-da.ua/', 'https://eva.ua/ua/', 'https://prostor.ua/ua/',
        'https://www.watsons.ua/']
for url in urls:
    r = requests.get(url, headers=header)
    # soup = BeautifulSoup(r.content, 'lxml')
    ws['A' + str(i)] = r.text
    i += 1
wb.save("sample.xlsx")
